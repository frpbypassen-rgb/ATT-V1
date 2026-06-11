import sys
import os
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, session, flash, Response, jsonify
from werkzeug.utils import secure_filename
import datetime
from bson import ObjectId
import io

# استدعاء مكتبات التنسيق للإكسيل
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 🌟 1. إجبار بايثون على قراءة المجلد الرئيسي (ATT_V2) أولاً
root_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, root_path)

# 🌟 2. الآن نستدعي الملفات بعد أن تعرف بايثون على المسار
from receipt_generator import create_image_receipt

try:
    from config import db, users, stock, transactions, invoices, ADMIN_IDS, customer_bot
except ImportError:
    print("🚨 خطأ: لم يتم العثور على ملف config.py في المجلد الرئيسي.")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'static')

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.secret_key = "AlAhram_Secret_Key_2026_Secure_Final_V14"

UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

ADMIN_CREDENTIALS = {
    "mohammed": "ahram2026",
    "admin": "admin123"
}

def is_admin():
    return session.get('logged_in') and session.get('username') in ADMIN_CREDENTIALS

# ==========================================
# 🌟 قسم الـ API الخاص بالعملاء والموزعين 🌟
# ==========================================
def check_api_auth():
    api_key = request.headers.get('x-api-key')
    return api_key == "AlAhram_Secret_Key_2026"

# 1. رابط API لجلب (منتج واحد محدد)
@app.route('/api/v1/product/<path:product_name>', methods=['GET'])
def api_get_single_product(product_name):
    if not check_api_auth():
        return jsonify({"error": "Unauthorized. Invalid API Key"}), 401
        
    clean_name = product_name.strip()
    available_cards = list(stock.find({"product": clean_name, "sold": False}))
    
    if not available_cards:
        return jsonify({"product": clean_name, "available": False, "message": "Product is out of stock"}), 404
        
    sample_card = available_cards[0]
    return jsonify({
        "product": sample_card.get('product'),
        "category": sample_card.get('category'),
        "price": float(sample_card.get('price_1', 0)),
        "available_quantity": len(available_cards),
        "available": True
    })

# 2. رابط API لجلب (قسم محدد)
@app.route('/api/v1/category/<path:category_name>', methods=['GET'])
def api_get_category_products(category_name):
    if not check_api_auth():
        return jsonify({"error": "Unauthorized"}), 401
        
    clean_category = category_name.strip()
    products = list(stock.find({"category": clean_category, "sold": False}, {"_id": 0, "product": 1, "price_1": 1}))
    unique_products = {}
    for item in products:
        prod_name = item.get('product')
        if prod_name and prod_name not in unique_products:
            unique_products[prod_name] = {"product": prod_name, "price": float(item.get('price_1', 0)), "available_quantity": 1}
        elif prod_name in unique_products:
            unique_products[prod_name]["available_quantity"] += 1
    return jsonify(list(unique_products.values()))

# 3. رابط API لجلب (جميع المنتجات)
@app.route('/api/v1/products', methods=['GET'])
def api_get_all_products():
    if not check_api_auth():
        return jsonify({"error": "Unauthorized"}), 401
    
    products = list(stock.find({"sold": False}, {"_id": 0, "product": 1, "category": 1, "price_1": 1}))
    unique_products = {}
    for item in products:
        prod_name = item.get('product')
        if prod_name and prod_name not in unique_products:
            unique_products[prod_name] = {"product": prod_name, "category": item.get('category'), "price": float(item.get('price_1', 0)), "available_quantity": 1}
        elif prod_name in unique_products:
            unique_products[prod_name]["available_quantity"] += 1
    return jsonify(list(unique_products.values()))

# 🚀 4. الرابط الجديد: شراء منتج وخصم الرصيد للعميل مباشرة 🚀
@app.route('/api/v1/buy', methods=['GET', 'POST'])
def api_buy_product():
    # يدعم قراءة بيانات العميل من الرابط المباشر (GET) أو عبر طلب برمجي (POST)
    if request.method == 'POST' and request.is_json:
        user_id = request.json.get('user_id')
        api_key = request.json.get('api_key')
        product_name = request.json.get('product')
    else:
        user_id = request.args.get('user_id')
        api_key = request.args.get('api_key')
        product_name = request.args.get('product')

    if not user_id or not api_key or not product_name:
        return jsonify({"error": "بيانات مفقودة. مطلوب: user_id, api_key, product"}), 400
        
    try: uid = int(user_id)
    except: uid = user_id
        
    # 1. التحقق من هوية العميل باستخدام كلمة مرور الموقع كـ (مفتاح API)
    user = users.find_one({"_id": uid, "web_password": str(api_key)})
    if not user:
        return jsonify({"error": "هوية غير مصرح بها. تأكد من رقم العميل وكلمة المرور."}), 401
        
    if not user.get('is_active', True):
        return jsonify({"error": "هذا الحساب موقوف من قبل الإدارة."}), 403

    clean_name = product_name.strip()
    
    # 2. البحث عن كرت متوفر
    card = stock.find_one({"product": clean_name, "sold": False})
    if not card:
        return jsonify({"error": "عذراً، هذا المنتج غير متوفر في المخزون حالياً."}), 404
        
    price = float(card.get('price_1', 0))
    balance = float(user.get('balance', 0))
    
    # 3. التحقق من الرصيد
    if balance < price:
        return jsonify({"error": "رصيدك غير كافٍ لإتمام العملية.", "balance": balance, "price": price}), 402
        
    # 4. إتمام المعاملة
    order_id = f"API-{datetime.datetime.now().strftime('%y%m%d%H%M%S')}"
    
    # تحديث الكرت كمباع
    stock.update_one(
        {"_id": card["_id"]}, 
        {"$set": {"sold": True, "order_id": order_id, "sold_date": datetime.datetime.now()}}
    )
    
    # خصم الرصيد
    users.update_one({"_id": uid}, {"$inc": {"balance": -price}})
    
    # تسجيل الفاتورة
    transactions.insert_one({
        "user_id": uid,
        "order_id": order_id,
        "product": clean_name,
        "qty": 1,
        "total_price": price,
        "date": datetime.datetime.now(),
        "type": "api_purchase"
    })
    
    # 5. تسليم الكرت للعميل
    return jsonify({
        "success": True,
        "message": "تم الشراء والخصم بنجاح",
        "product": clean_name,
        "price": price,
        "remaining_balance": balance - price,
        "order_id": order_id,
        "card_details": {
            "code": card.get("code", "---"),
            "serial": card.get("serial", "---"),
            "pin": card.get("pin", "---"),
            "op_code": card.get("op_code", "---")
        }
    })
# ==========================================


def get_customer_name(user_id):
    if not user_id: return "مجهول"
    uid_str = str(user_id).strip()
    try: uid_int = int(uid_str)
    except: uid_int = None
    
    search_queries = [{"_id": uid_int}, {"_id": uid_str}, {"user_id": uid_int}, {"user_id": uid_str}]
    valid_queries = [q for q in search_queries if list(q.values())[0] is not None]
    
    try:
        u_data = users.find_one({"$or": valid_queries})
    except: u_data = None
        
    if u_data:
        name = u_data.get('name') or u_data.get('first_name') or u_data.get('username')
        if name: return str(name).strip()
    return uid_str

@app.route('/')
def index():
    if not is_admin(): return redirect(url_for('login'))
    total_users = users.count_documents({})
    available_stock = stock.count_documents({"sold": False})
    sold_count = stock.count_documents({"sold": True})
    
    pipeline_total = [{"$group": {"_id": None, "total": {"$sum": "$total_price"}}}]
    sales_data = list(transactions.aggregate(pipeline_total))
    sales_val = sales_data[0]['total'] if sales_data else 0
    formatted_sales = "{:.2f}".format(float(sales_val if sales_val is not None else 0))

    today = datetime.datetime.now()
    first_day = datetime.datetime(today.year, today.month, 1)
    pipeline_chart = [
        {"$match": {"date": {"$gte": first_day}}},
        {"$group": {"_id": "$user_id", "total_spent": {"$sum": "$total_price"}}},
        {"$sort": {"total_spent": -1}}, {"$limit": 10}
    ]
    chart_results = list(transactions.aggregate(pipeline_chart))
    chart_labels = [get_customer_name(item['_id']) for item in chart_results]
    chart_values = [round(item['total_spent'], 2) for item in chart_results]

    return render_template('index.html', users_count=total_users, stock=available_stock, 
                           sold_count=sold_count, sales=formatted_sales,
                           chart_labels=chart_labels, chart_values=chart_values)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username in ADMIN_CREDENTIALS and ADMIN_CREDENTIALS[username] == password:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index'))
        else: flash("بيانات الدخول غير صحيحة!", "danger")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/users')
def manage_users():
    if not is_admin(): return redirect(url_for('login'))
    all_users = list(users.find().sort("balance", -1))
    for u in all_users:
        u['display_name'] = u.get('name') or u.get('first_name') or u.get('username') or u.get('_id')
        u['user_id_field'] = u.get('_id')
        balance_val = u.get('balance', 0)
        u['formatted_balance'] = "{:.2f}".format(float(balance_val if balance_val is not None else 0))
    return render_template('users.html', customers=all_users)

@app.route('/update_balance/<user_id>', methods=['POST'])
def update_balance(user_id):
    if not is_admin(): return redirect(url_for('login'))
    new_balance = request.form.get('new_balance', type=float)
    redirect_url = request.form.get('redirect_url', url_for('manage_users'))
    
    if new_balance is not None:
        try: uid = int(user_id)
        except: uid = user_id
        
        users.update_one({"_id": uid}, {"$set": {"balance": new_balance}})
        
        try:
            msg_text = f"⚙️ **تحديث إداري للرصيد**\n\nتم ضبط رصيدك الكلي في المنظومة ليصبح:\n💰 `{new_balance:.2f} د.ل`"
            customer_bot.send_message(uid, msg_text, parse_mode="Markdown")
        except: pass
            
        flash("تم ضبط الرصيد وإخطار العميل بنجاح!", "success")
        
    return redirect(redirect_url)

@app.route('/charge_balance/<user_id>', methods=['POST'])
def charge_balance(user_id):
    if not is_admin(): return redirect(url_for('login'))
    amount = request.form.get('amount', type=float)
    if amount is not None:
        try: uid = int(user_id)
        except: uid = user_id
        
        operation_time = datetime.datetime.now()
        user_info = users.find_one({"_id": uid})
        prev_balance = float(user_info.get('balance', 0))
        new_total = prev_balance + amount
        user_name = user_info.get('name', 'عميل الأهرام')
        user_phone = user_info.get('phone', '---')
        
        users.update_one({"_id": uid}, {"$inc": {"balance": amount}})
        
        try:
            image_path = create_image_receipt(
                name=user_name, phone=user_phone, user_id=uid, 
                amount=amount, prev_balance=prev_balance, total_balance=new_total,
                op_date=operation_time
            )
            
            if image_path and os.path.exists(image_path):
                with open(image_path, 'rb') as photo:
                    caption = f"✅ **تم إضافة رصيد لحسابك بنجاح!**\n📅 التاريخ: `{operation_time.strftime('%Y-%m-%d %I:%M %p')}`"
                    customer_bot.send_photo(uid, photo, caption=caption, parse_mode="Markdown")
                os.remove(image_path)
            else:
                customer_bot.send_message(uid, f"✅ **تم شحن رصيدك!**\nالقيمة المضافة: `{amount:.2f}` د.ل\nالرصيد الكلي: `{new_total:.2f}` د.ل", parse_mode="Markdown")
        except Exception as e: 
            print(f"Error sending image receipt: {e}")
            
        flash(f"تم شحن {amount} د.ل وإرسال الإيصال المصور للعميل!", "success")
    return redirect(url_for('user_profile', user_id=user_id))

@app.route('/update_user_info/<user_id>', methods=['POST'])
def update_user_info(user_id):
    if not is_admin(): return redirect(url_for('login'))
    new_name = request.form.get('name', '').strip()
    new_phone = request.form.get('phone', '').strip()
    
    try: uid = int(user_id)
    except: uid = user_id
    
    update_data = {}
    if new_name: update_data['name'] = new_name
    if new_phone: update_data['phone'] = new_phone
    
    if update_data:
        users.update_one({"_id": uid}, {"$set": update_data})
        flash("تم تحديث بيانات العميل بنجاح!", "success")
    return redirect(url_for('user_profile', user_id=user_id))

@app.route('/toggle_user_status/<user_id>', methods=['POST'])
def toggle_user_status(user_id):
    if not is_admin(): return redirect(url_for('login'))
    try: uid = int(user_id)
    except: uid = user_id
    
    user = users.find_one({"_id": uid})
    if user:
        current_status = user.get('is_active', True)
        new_status = not current_status
        users.update_one({"_id": uid}, {"$set": {"is_active": new_status}})
        
        try:
            if new_status:
                msg_text = "✅ **تم إعادة تنشيط حسابك**\nيمكنك الآن استخدام خدمات المنظومة مجدداً."
            else:
                msg_text = "🚫 **تم إيقاف حسابك مؤقتاً**\nيرجى التواصل مع الإدارة للمزيد من التفاصيل."
            customer_bot.send_message(uid, msg_text, parse_mode="Markdown")
        except: pass
            
        action = "تنشيط" if new_status else "إيقاف"
        flash(f"تم {action} الحساب وإخطار العميل!", "success")
        
    return redirect(url_for('user_profile', user_id=user_id))

@app.route('/user_profile/<user_id>')
def user_profile(user_id):
    if not is_admin(): return redirect(url_for('login'))
    try: uid = int(user_id)
    except: uid = user_id
    
    search_queries = [{"_id": uid}, {"_id": str(uid)}, {"user_id": uid}, {"user_id": str(uid)}]
    user_data = users.find_one({"$or": search_queries})
    
    if not user_data:
        flash("هذا العميل غير موجود.", "danger")
        return redirect(url_for('manage_users'))
        
    display_name = user_data.get('name') or user_data.get('first_name') or user_data.get('username') or str(uid)
    formatted_balance = "{:.2f}".format(float(user_data.get('balance', 0)))
    
    user_data['user_id_field'] = user_data.get('_id', uid)
    user_data['username_field'] = user_data.get('username', 'لا يوجد')
    user_data['phone_field'] = user_data.get('phone', 'لا يوجد')
    user_data['is_active'] = user_data.get('is_active', True)
    
    user_trans = list(transactions.find({"user_id": {"$in": [uid, str(uid)]}}).sort("date", -1))
    invoices_dict = {}
    total_spent = 0
    
    for t in user_trans:
        oid = t.get('order_id')
        if not oid: continue
        price = t.get('total_price', 0)
        total_spent += price
        
        if oid not in invoices_dict:
            cards_count = stock.count_documents({"order_id": oid})
            invoices_dict[oid] = {
                "order_id": oid, "total_price": 0, "date": t.get('date'),
                "cards_count": cards_count if cards_count > 0 else 1
            }
        invoices_dict[oid]["total_price"] += price
        
    enriched_invoices = []
    for oid, data in invoices_dict.items():
        d_obj = data["date"]
        data["formatted_price"] = "{:.2f}".format(float(data["total_price"]))
        data["date_str"] = d_obj.strftime('%Y-%m-%d') if isinstance(d_obj, datetime.datetime) else "---"
        data["time_str"] = d_obj.strftime('%H:%M:%S') if isinstance(d_obj, datetime.datetime) else "---"
        enriched_invoices.append(data)
        
    enriched_invoices.sort(key=lambda x: x.get("date") if isinstance(x.get("date"), datetime.datetime) else datetime.datetime.min, reverse=True)
    formatted_total_spent = "{:.2f}".format(float(total_spent))
    
    return render_template('user_profile.html', user=user_data, display_name=display_name, balance=formatted_balance, total_spent=formatted_total_spent, invoices=enriched_invoices, invoices_count=len(enriched_invoices))

@app.route('/search_card', methods=['GET', 'POST'])
def search_card():
    if not is_admin(): return redirect(url_for('login'))
    search_result = None; error_msg = None
    if request.method == 'POST':
        query = request.form.get('query', '').strip()
        if query:
            clean_query = query.replace('.0', '')
            card = stock.find_one({"$or": [{"code": clean_query}, {"serial": clean_query}, {"code": query}, {"serial": query}]})
            if card:
                price_val = card.get("price_1", 0)
                formatted_price = "{:.2f}".format(float(price_val if price_val is not None else 0))
                search_result = {"product": card.get("product", "---"), "category": card.get("category", "---"), "code": card.get("code", "---"), "serial": card.get("serial", "---"), "price": formatted_price, "sold": card.get("sold", False), "added_date": card.get("added_date", "---")}
                if isinstance(search_result["added_date"], datetime.datetime): search_result["added_date"] = search_result["added_date"].strftime('%Y-%m-%d %H:%M')
                if search_result["sold"]:
                    order_id = card.get("order_id")
                    search_result["order_id"] = order_id
                    trans = transactions.find_one({"order_id": order_id})
                    if trans:
                        search_result["customer_name"] = get_customer_name(trans.get("user_id"))
                        sell_date = trans.get("date")
                        search_result["sell_date"] = sell_date.strftime('%Y-%m-%d %H:%M') if isinstance(sell_date, datetime.datetime) else "---"
                    else: search_result["customer_name"] = "مجهول"; search_result["sell_date"] = "---"
            else: error_msg = "لم يتم العثور على كرت يطابق البحث."
    return render_template('search_card.html', result=search_result, error=error_msg)

@app.route('/invoices')
def view_invoices():
    if not is_admin(): return redirect(url_for('login'))
    all_transactions = list(transactions.find().sort("date", -1))
    enriched_invoices = []
    seen_orders = set()
    for trans in all_transactions:
        order_id = trans.get('order_id')
        if not order_id or order_id in seen_orders: continue
        seen_orders.add(order_id)
        customer_name = get_customer_name(trans.get('user_id'))
        cards_count = stock.count_documents({"order_id": order_id})
        date_obj = trans.get('date')
        order_all_trans = [t for t in all_transactions if t.get('order_id') == order_id]
        order_total_price = sum(t.get('total_price', 0) for t in order_all_trans)
        enriched_invoices.append({
            "order_id": order_id, "customer_name": customer_name,
            "total_price": "{:.2f}".format(float(order_total_price)),
            "cards_count": cards_count if cards_count > 0 else 1,
            "date": date_obj.strftime('%Y-%m-%d') if isinstance(date_obj, datetime.datetime) else "---",
            "time": date_obj.strftime('%H:%M:%S') if isinstance(date_obj, datetime.datetime) else "---"
        })
    return render_template('invoices.html', invoices=enriched_invoices)

@app.route('/download_invoice/<order_id>')
def download_invoice(order_id):
    if not is_admin(): return redirect(url_for('login'))
    all_trans = list(transactions.find({"order_id": order_id}))
    if not all_trans: return "الفاتورة غير موجودة", 404
    cards = list(stock.find({"order_id": order_id}))
    customer_name = get_customer_name(all_trans[0].get('user_id'))
    order_date = all_trans[0].get('date')
    date_str = order_date.strftime('%Y-%m-%d %H:%M') if isinstance(order_date, datetime.datetime) else "---"

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_list = []
        for t in all_trans:
            qty = t.get('qty', 1)
            total = t.get('total_price', 0)
            unit_price = total / qty if qty > 0 else 0
            summary_list.append({'اسم المنتج': t.get('product'), 'العدد': qty, 'السعر': unit_price, 'المجموع': total})
        df_sum = pd.DataFrame(summary_list)
        df_sum.insert(0, 'ت', range(1, 1 + len(df_sum)))
        df_sum.to_excel(writer, sheet_name="📊 ملخص الفاتورة", index=False, startrow=6)
        
        products = set([c.get('product', 'Unknown') for c in cards])
        for prod in products:
            prod_cards = [c for c in cards if c.get('product') == prod]
            df = pd.DataFrame(prod_cards)
            expected_cols = ['product', 'code', 'serial', 'pin', 'op_code']
            for col in expected_cols:
                if col not in df.columns: df[col] = pd.NA
            for col in ['code', 'pin', 'serial', 'op_code']:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)
                    df[col] = df[col].replace(['nan', 'None', 'NaN', ''], '-')
            df = df[expected_cols]
            df.rename(columns={'product': 'المنتج', 'code': 'الكود', 'serial': 'السيريال', 'pin': 'الرقم السري', 'op_code': 'أوبريشن'}, inplace=True)
            df.insert(0, 'ت', range(1, 1 + len(df)))
            safe_sheet_name = "".join([c for c in prod if c.isalnum() or c in " _-"])[:30]
            if not safe_sheet_name: safe_sheet_name = "مشتريات"
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False, startrow=6)

        workbook = writer.book
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        password_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        footer_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=14)
        bold_font = Font(bold=True, size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal="center", vertical="center")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.sheet_view.rightToLeft = True
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A5
            last_col_letter = get_column_letter(worksheet.max_column)

            worksheet.merge_cells(f'A1:{last_col_letter}1')
            worksheet["A1"] = "🏢 شركة الأهرام للاتصالات والتقنية"
            worksheet["A1"].font, worksheet["A1"].fill, worksheet["A1"].alignment = white_font, header_fill, align
            worksheet.row_dimensions[1].height = 30

            worksheet.merge_cells(f'A2:{last_col_letter}2'); worksheet["A2"] = "تفاصيل الفاتورة"
            worksheet["A2"].font, worksheet["A2"].alignment = Font(bold=True, size=12), align

            worksheet.merge_cells(f'A3:{last_col_letter}3'); worksheet["A3"] = f"رقم الفاتورة: {order_id}"
            worksheet["A3"].alignment = align

            worksheet.merge_cells(f'A4:{last_col_letter}4'); worksheet["A4"] = f"اسم العميل: {customer_name}"
            worksheet["A4"].alignment, worksheet["A4"].font = align, Font(bold=True, color="1F4E78", size=12)

            worksheet.merge_cells(f'A5:{last_col_letter}5'); worksheet["A5"] = f"التاريخ: {date_str}"
            worksheet["A5"].alignment = align

            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=7, column=col_num)
                cell.fill, cell.font, cell.alignment, cell.border = header_fill, Font(color="FFFFFF", bold=True), align, thin_border

            for row_idx in range(8, worksheet.max_row + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment, cell.border, cell.font = align, thin_border, bold_font
                    cell.number_format = '@'
                    if cell.value is not None and str(cell.value) != '-':
                        cell.value = str(cell.value)
                    if "ملخص" not in sheet_name and worksheet.cell(row=7, column=col_idx).value == 'الرقم السري':
                        cell.fill = password_fill

            for i, col in enumerate(worksheet.columns, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = 20

            if "ملخص" in sheet_name:
                last_r = worksheet.max_row + 2
                grand_total = sum(d['المجموع'] for d in summary_list)
                total_qty = sum(d['العدد'] for d in summary_list)
                
                worksheet.cell(row=last_r, column=2, value="العدد الكلي:").fill = footer_fill
                worksheet.cell(row=last_r, column=2).font, worksheet.cell(row=last_r, column=2).alignment = bold_font, align
                worksheet.cell(row=last_r, column=3, value=total_qty).fill = footer_fill
                worksheet.cell(row=last_r, column=3).font, worksheet.cell(row=last_r, column=3).alignment = bold_font, align
                worksheet.cell(row=last_r, column=4, value="المبلغ الكلي:").fill = footer_fill
                worksheet.cell(row=last_r, column=4).font, worksheet.cell(row=last_r, column=4).alignment = bold_font, align
                worksheet.cell(row=last_r, column=5, value=f"{grand_total:.2f} د.ل").fill = footer_fill
                worksheet.cell(row=last_r, column=5).font, worksheet.cell(row=last_r, column=5).alignment = bold_font, align
                
                thanks_row = last_r + 2
                worksheet.merge_cells(start_row=thanks_row, start_column=1, end_row=thanks_row+1, end_column=worksheet.max_column)
                thanks_cell = worksheet.cell(row=thanks_row, column=1)
                thanks_cell.value = "💖 شكراً لثقتكم وتسوقكم من شركة الأهرام للاتصالات والتقنية.. نتمنى لكم يوماً سعيداً!"
                thanks_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thanks_cell.font, thanks_cell.fill = Font(color="1F4E78", bold=True, size=13, italic=True), PatternFill(start_color="EAEEF7", end_color="EAEEF7", fill_type="solid")

    output.seek(0)
    return Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-disposition": f"attachment; filename=Invoice_{order_id}.xlsx"})

@app.route('/api/get_products', methods=['POST'])
def get_products_api_internal():
    if not is_admin(): return jsonify([])
    category = request.get_json().get('category')
    products_in_cat = stock.distinct("product", {"category": category})
    return jsonify(products_in_cat)

@app.route('/api/get_price', methods=['POST'])
def get_price():
    if not is_admin(): return jsonify({'price': ''})
    product_name = request.get_json().get('product')
    last_item = stock.find_one({"product": product_name}, sort=[("added_date", -1)])
    if last_item and 'price_1' in last_item:
        formatted_api_price = "{:.2f}".format(float(last_item['price_1']))
        return jsonify({'price': formatted_api_price})
    return jsonify({'price': ''})

@app.route('/download_template')
def download_template():
    if not is_admin(): return redirect(url_for('login'))
    output = io.BytesIO()
    df = pd.DataFrame(columns=['الكود', 'السيريال', 'الرقم السري', 'أوبريشن'])
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Template')
    output.seek(0)
    return Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-disposition": "attachment; filename=AlAhram_Template.xlsx"})

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if not is_admin(): return redirect(url_for('login'))
    if request.method == 'POST':
        category = request.form.get('category')
        product = request.form.get('product')
        price_1 = request.form.get('price_1', type=float)
        file = request.files.get('file')
        if file and (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            try:
                df = pd.read_csv(filepath) if filename.endswith('.csv') else pd.read_excel(filepath)
                code_col = next((c for c in df.columns if any(x in str(col).lower() for x in ['pin', 'code', 'كود', 'رقم'])), None)
                serial_col = next((c for c in df.columns if any(x in str(col).lower() for x in ['serial', 'سيريال'])), None)
                pin_col = next((c for c in df.columns if any(x in str(c).lower() for x in ['سري', 'pass', 'password']) and c != code_col), None)
                op_col = next((c for c in df.columns if any(x in str(c).lower() for x in ['op', 'اوب', 'أوب'])), None)
                if not code_col:
                    flash("لم يتم العثور على عمود الأكواد (الكود) في الملف!", "danger")
                    return redirect(url_for('upload_file'))
                codes = []
                for _, row in df.iterrows():
                    val = str(row[code_col]).strip()
                    if val and val.lower() != 'nan':
                        doc = {"category": category, "product": product, "price_1": price_1, "code": val.replace('.0', ''), "sold": False, "added_date": datetime.datetime.now()}
                        if serial_col and str(row[serial_col]).lower() != 'nan': doc["serial"] = str(row[serial_col]).strip().replace('.0', '')
                        if pin_col and str(row[pin_col]).lower() != 'nan': doc["pin"] = str(row[pin_col]).strip().replace('.0', '')
                        if op_col and str(row[op_col]).lower() != 'nan': doc["op_code"] = str(row[op_col]).strip().replace('.0', '')
                        codes.append(doc)
                if codes: stock.insert_many(codes); flash(f"تم إضافة {len(codes)} كرت بنجاح!", "success")
            except Exception as e: flash(f"حدث خطأ في قراءة الملف: {str(e)}", "danger")
            return redirect(url_for('upload_file'))
    categories = stock.distinct("category")
    return render_template('upload.html', categories=categories)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
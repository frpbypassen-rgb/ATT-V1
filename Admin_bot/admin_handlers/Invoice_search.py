import pandas as pd
import io
import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from config import admin_bot, transactions, stock, users
from telebot import types

# ==========================================
# 1. دالة بناء الإكسيل للإدارة المجمعة (محصنة بنظام الـ Text)
# ==========================================
def generate_admin_full_invoice(cards, all_trans, invoice_id, customer_name):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # --- [1] شيت الملخص ---
        summary_list = []
        for t in all_trans:
            qty = t.get('qty', 1)
            total = t.get('total_price', 0)
            unit_price = total / qty if qty > 0 else 0
            
            summary_list.append({
                'اسم المنتج': t.get('product'),
                'العدد': qty,
                'السعر': unit_price,
                'المجموع': total
            })
            
        df_sum = pd.DataFrame(summary_list)
        df_sum.insert(0, 'ت', range(1, 1 + len(df_sum)))
        df_sum.to_excel(writer, sheet_name="📊 ملخص الفاتورة", index=False, startrow=6)
        
        # --- [2] شيتات الأكواد ---
        products = set([c.get('product', 'Unknown') for c in cards])
        for prod in products:
            prod_cards = [c for c in cards if c.get('product') == prod]
            df = pd.DataFrame(prod_cards)
            
            expected_cols = ['product', 'code', 'serial', 'pin', 'op_code']
            for col in expected_cols:
                if col not in df.columns: df[col] = pd.NA
                    
            # 🌟 الحماية 1: تحويل السيريال والأكواد إلى نصوص وإزالة .0
            for col in ['code', 'pin', 'serial', 'op_code']:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)
                    df[col] = df[col].replace(['nan', 'None', 'NaN', ''], '-')

            df = df[expected_cols]
            df.rename(columns={'product': 'المنتج', 'code': 'الكود', 'serial': 'السيريال', 'pin': 'الرقم السري', 'op_code': 'أوبريشن'}, inplace=True)
            df.insert(0, 'ت', range(1, 1 + len(df)))

            safe_sheet_name = "".join([c for c in prod if c.isalnum() or c in " _-"])[:30]
            if not safe_sheet_name: safe_sheet_name = "مشتريات"
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False, startrow=6)

        # --- [3] التنسيقات المطابقة لفاتورة العميل ---
        workbook = writer.book
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        password_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        footer_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=14)
        bold_font = Font(bold=True, size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal="center", vertical="center")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.sheet_view.rightToLeft = True
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A5
            
            num_cols = worksheet.max_column
            last_col_letter = get_column_letter(num_cols)

            worksheet.merge_cells(f'A1:{last_col_letter}1')
            worksheet["A1"] = "🏢 شركة الأهرام للاتصالات والتقنية"
            worksheet["A1"].font = white_font
            worksheet["A1"].fill = header_fill
            worksheet["A1"].alignment = align
            worksheet.row_dimensions[1].height = 30

            worksheet.merge_cells(f'A2:{last_col_letter}2')
            worksheet["A2"] = f"تفاصيل الفاتورة"
            worksheet["A2"].font = Font(bold=True, size=12)
            worksheet["A2"].alignment = align

            worksheet.merge_cells(f'A3:{last_col_letter}3')
            worksheet["A3"] = f"رقم الفاتورة: {invoice_id}"
            worksheet["A3"].alignment = align

            worksheet.merge_cells(f'A4:{last_col_letter}4')
            worksheet["A4"] = f"اسم العميل: {customer_name}"
            worksheet["A4"].alignment = align
            worksheet["A4"].font = Font(bold=True, color="1F4E78", size=12)

            worksheet.merge_cells(f'A5:{last_col_letter}5')
            worksheet["A5"] = f"التاريخ: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
            worksheet["A5"].alignment = align

            for col_num in range(1, num_cols + 1):
                cell = worksheet.cell(row=7, column=col_num)
                cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = align; cell.border = thin_border

            for row_idx in range(8, worksheet.max_row + 1):
                for col_idx in range(1, num_cols + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = align; cell.border = thin_border; cell.font = bold_font
                    
                    # 🌟 الحماية 2: إجبار الإكسيل على قراءة الخلية كنص (Text)
                    cell.number_format = '@'
                    
                    if cell.value is not None and str(cell.value) != '-':
                        cell.value = str(cell.value)
                    
                    if "ملخص" not in sheet_name and worksheet.cell(row=7, column=col_idx).value == 'الرقم السري':
                        cell.fill = password_fill

            for i, col in enumerate(worksheet.columns, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = 20

            if "ملخص" in sheet_name:
                last_r = worksheet.max_row + 2
                grand_total = sum(d['المجموع'] for d in summary_list)
                total_qty = sum(d['العدد'] for d in summary_list)
                
                worksheet.cell(row=last_r, column=2, value="العدد الكلي:").fill = footer_fill
                worksheet.cell(row=last_r, column=2).font = bold_font; worksheet.cell(row=last_r, column=2).alignment = align
                
                worksheet.cell(row=last_r, column=3, value=total_qty).fill = footer_fill
                worksheet.cell(row=last_r, column=3).font = bold_font; worksheet.cell(row=last_r, column=3).alignment = align
                
                worksheet.cell(row=last_r, column=4, value="المبلغ الكلي:").fill = footer_fill
                worksheet.cell(row=last_r, column=4).font = bold_font; worksheet.cell(row=last_r, column=4).alignment = align
                
                worksheet.cell(row=last_r, column=5, value=f"{grand_total:.2f} د.ل").fill = footer_fill
                worksheet.cell(row=last_r, column=5).font = bold_font; worksheet.cell(row=last_r, column=5).alignment = align
                
                thanks_row = last_r + 2
                worksheet.merge_cells(start_row=thanks_row, start_column=1, end_row=thanks_row+1, end_column=num_cols)
                thanks_cell = worksheet.cell(row=thanks_row, column=1)
                thanks_cell.value = "💖 شكراً لثقتكم وتسوقكم من شركة الأهرام للاتصالات والتقنية.. نتمنى لكم يوماً سعيداً!"
                thanks_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thanks_cell.font = Font(color="1F4E78", bold=True, size=13, italic=True)
                thanks_cell.fill = PatternFill(start_color="EAEEF7", end_color="EAEEF7", fill_type="solid")

    output.seek(0)
    return output

# ==========================================
# 2. الواجهة وطلب المدخلات
# ==========================================
@admin_bot.message_handler(func=lambda m: m.text == "🔍 البحث عن فاتورة")
def start_search(msg):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    kb.add("🔙 إلغاء البحث")
    
    text = (
        "🔍 **قسم البحث الشامل:**\n\n"
        "يمكنك إدخال أي من البيانات التالية:\n"
        "🔹 **رقم الفاتورة** (مثال: `ATT-2604-001`)\n"
        "🔹 **رقم الـ ID للعميل**\n"
        "🔹 **رقم هاتف العميل**"
    )
    res = admin_bot.send_message(msg.chat.id, text, parse_mode="Markdown", reply_markup=kb)
    admin_bot.register_next_step_handler(res, process_hybrid_search)

# ==========================================
# 3. محرك البحث الهجين (فاتورة أو عميل)
# ==========================================
def process_hybrid_search(msg):
    uid = msg.chat.id
    search_query = msg.text.strip()

    if search_query == "🔙 إلغاء البحث":
        admin_bot.clear_step_handler_by_chat_id(uid)
        return admin_bot.send_message(uid, "🔙 تم العودة للقائمة الرئيسية.")

    # 🌟 [أ] محاولة البحث كـ رقم فاتورة أولاً (إذا كان يحتوي على ATT أو INV)
    trans_query = {"order_id": {"$regex": search_query, "$options": "i"}}
    invoice_match = list(transactions.find(trans_query))
    
    if invoice_match:
        actual_id = invoice_match[0]['order_id']
        all_trans = list(transactions.find({"order_id": actual_id}))
        return send_invoice_file(uid, actual_id, all_trans)

    # 🌟 [ب] إذا لم تكن فاتورة، سنجرب البحث كعميل (آيدي أو رقم هاتف)
    search_vars = [search_query, str(search_query)]
    try: search_vars.append(int(search_query))
    except: pass

    # البحث في جدول المستخدمين
    user = users.find_one({
        "$or": [
            {"_id": {"$in": search_vars}},
            {"phone": {"$in": search_vars}}
        ]
    })

    if user:
        # وجدنا العميل! نأخذ الآيدي الخاص به لجلب الفواتيره
        target_ids = [user['_id'], str(user['_id'])]
        try: target_ids.append(int(user['_id']))
        except: pass
        customer_name = user.get('name', 'عميل مسجل')
    else:
        # ربما لم يجده في جدول المستخدمين، نجرب البحث في المبيعات مباشرة بالمدخل
        target_ids = search_vars
        customer_name = "عميل (غير مسجل بالاسم)"

    user_orders = transactions.distinct("order_id", {"user_id": {"$in": target_ids}})

    if user_orders:
        kb = types.InlineKeyboardMarkup(row_width=1)
        # عرض آخر الفواتير من الأحدث للأقدم
        for order in reversed(user_orders): 
            kb.add(types.InlineKeyboardButton(f"📄 {order}", callback_data=f"pull_inv_{order}"))
        
        text = f"👤 **تم العثور على {len(user_orders)} فواتير للعميل:** `{customer_name}`\n\nتفضل باختيار الفاتورة المطلوبة:"
        return admin_bot.send_message(uid, text, reply_markup=kb, parse_mode="Markdown")

    # 🌟 [ج] إذا فشل كل شيء
    admin_bot.send_message(uid, "❌ **لم يتم العثور على أي بيانات!**\nتأكد من إدخال رقم الفاتورة أو رقم/ID العميل بشكل صحيح.", parse_mode="Markdown")

# ==========================================
# 4. دالة مساعدة لإرسال الفاتورة
# ==========================================
def send_invoice_file(chat_id, order_id, all_trans, call_id=None):
    if call_id:
        admin_bot.answer_callback_query(call_id, "⏳ جاري استخراج الفاتورة الشاملة...")
    else:
        admin_bot.send_message(chat_id, "⏳ جاري استخراج الفاتورة الشاملة...")
        
    all_cards = list(stock.find({"order_id": order_id}))
    
    try:
        user_id = all_trans[0].get('user_id')
        user = users.find_one({"_id": user_id})
        customer_name = user.get('name', 'غير مسجل') if user else str(user_id)

        excel_file = generate_admin_full_invoice(all_cards, all_trans, order_id, customer_name)
        total_val = sum(t.get('total_price', 0) for t in all_trans)
        
        summary = (
            f"🧾 **فاتورة مجمعة مستخرجة:**\n"
            f"━━━━━━━━━━━━━━━\n"
            f"👤 **العميل:** {customer_name}\n"
            f"💰 **الإجمالي:** `{total_val:.2f}` د.ل\n"
            f"━━━━━━━━━━━━━━━"
        )
        admin_bot.send_document(chat_id, excel_file, visible_file_name=f"{order_id}_Archive.xlsx", caption=summary, parse_mode="Markdown")
    except Exception as e:
        print(f"🚨 خطأ: {e}")
        admin_bot.send_message(chat_id, f"❌ حدث خطأ داخلي أثناء التوليد.")

@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("pull_inv_"))
def pull_and_send_invoice(call):
    order_id = call.data.replace("pull_inv_", "")
    all_trans = list(transactions.find({"order_id": order_id}))
    
    if not all_trans:
        return admin_bot.answer_callback_query(call.id, "❌ خطأ: الفاتورة غير موجودة.", show_alert=True)
    
    send_invoice_file(call.message.chat.id, order_id, all_trans, call.id)
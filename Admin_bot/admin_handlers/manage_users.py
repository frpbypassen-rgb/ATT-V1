import datetime
import pandas as pd
import io
from config import admin_bot, users, transactions, db, customer_bot
from telebot import types
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter 

# ==========================================
# 1. تفعيل العميل من الإشعار وعرض بطاقة التحكم
# ==========================================
@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("set_status_active_"))
def admin_activate_user_callback(call):
    user_id = int(call.data.replace("set_status_active_", ""))
    
    # استخدام البحث المرن لتفادي مشكلة (نص/رقم)
    result = users.update_one({"_id": {"$in": [user_id, str(user_id)]}}, {"$set": {"status": "active"}})
    
    if result.modified_count > 0 or users.find_one({"_id": {"$in": [user_id, str(user_id)]}, "status": "active"}):
        try:
            customer_bot.send_message(user_id, "🎉 **تهانينا! تم تفعيل حسابك بنجاح.**\nيمكنك الآن إرسال /start للبدء في استخدام المتجر وشراء الكروت.", parse_mode="Markdown")
        except: pass

        admin_bot.answer_callback_query(call.id, "✅ تم تفعيل العميل بنجاح")
        show_user_control_card(call.message.chat.id, user_id, call.message.message_id)
    else:
        admin_bot.answer_callback_query(call.id, "⚠️ العميل غير موجود.", show_alert=True)

# ==========================================
# 2. دالة بناء "بطاقة العميل" الاحترافية
# ==========================================
def show_user_control_card(admin_chat_id, user_id, message_id_to_edit=None):
    user = users.find_one({"_id": {"$in": [user_id, str(user_id)]}})
    if not user:
        return admin_bot.send_message(admin_chat_id, "❌ تعذر جلب بيانات العميل.")

    level_names = {1: "قطاعي عادي", 2: "موزع", 3: "جملة"}
    user_level = level_names.get(user.get('level', 1), "غير محدد")
    
    # سحب حد الائتمان السالب (الافتراضي صفر)
    credit_limit = user.get('credit_limit', 0.0)
    
    status_db = user.get('status', 'pending')
    if status_db == 'active': status_text = "🟢 مفعّل"
    elif status_db == 'blocked': status_text = "🔴 محظور"
    else: status_text = "⏳ معلق"
    
    card_text = (
        f"👤 **بطاقة بيانات العميل**\n"
        f"━━━━━━━━━━━━━━━\n"
        f"📝 **الاسم:** {user.get('name')}\n"
        f"📞 **الهاتف:** `{user.get('phone')}`\n"
        f"🆔 **ID:** `{user.get('_id')}`\n"
        f"💰 **الرصيد الحالي:** `{user.get('balance', 0.0):.2f}` د.ل\n"
        f"🛑 **حد الائتمان (السالب):** `{credit_limit:.2f}` د.ل\n"
        f"📊 **المستوى:** {user_level}\n"
        f"الحالة: {status_text}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"🎮 **أدوات التحكم السريع:**"
    )

    block_btn_text = "✅ فك الحظر" if status_db == 'blocked' else "🚫 حظر العميل"
    block_btn_callback = f"unblock_user_{user_id}" if status_db == 'blocked' else f"block_user_{user_id}"

    kb = types.InlineKeyboardMarkup(row_width=2)
    kb.add(
        types.InlineKeyboardButton("🏷️ تغيير المستوى", callback_data=f"change_level_{user_id}"),
        types.InlineKeyboardButton("💰 شحن رصيد", callback_data=f"admin_charge_{user_id}")
    )
    kb.add(
        types.InlineKeyboardButton("🛑 ضبط حد السالب", callback_data=f"set_credit_{user_id}"),
        types.InlineKeyboardButton(block_btn_text, callback_data=block_btn_callback)
    )
    kb.add(types.InlineKeyboardButton("🔙 إغلاق البطاقة", callback_data="close_card"))

    if message_id_to_edit:
        try: admin_bot.edit_message_text(card_text, admin_chat_id, message_id_to_edit, reply_markup=kb, parse_mode="Markdown")
        except: pass
    else:
        admin_bot.send_message(admin_chat_id, card_text, reply_markup=kb, parse_mode="Markdown")

@admin_bot.callback_query_handler(func=lambda call: call.data == "close_card")
def close_admin_card(call):
    try: admin_bot.delete_message(call.message.chat.id, call.message.message_id)
    except: pass

# ==========================================
# 3. برمجة "حد الائتمان الائتماني" (السالب)
# ==========================================
@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("set_credit_"))
def admin_set_credit_prompt(call):
    user_id = int(call.data.replace("set_credit_", ""))
    admin_bot.answer_callback_query(call.id)
    res = admin_bot.send_message(call.message.chat.id, "🛑 **ضبط حد الائتمان:**\nأرسل أقصى مبلغ يمكن للعميل سحبه بالسالب (مثلاً: 50 أو 100):")
    admin_bot.register_next_step_handler(res, process_set_credit, user_id)

def process_set_credit(msg, user_id):
    try:
        limit = float(msg.text.strip())
        if limit < 0:
            limit = abs(limit) # في حال أدخل المدير الرقم بالسالب بالخطأ
            
        users.update_one({"_id": {"$in": [user_id, str(user_id)]}}, {"$set": {"credit_limit": limit}})
        admin_bot.send_message(msg.chat.id, f"✅ تم تحديد حد الائتمان للعميل بقيمة `{limit:.2f}` د.ل")
        show_user_control_card(msg.chat.id, user_id)
    except ValueError:
        admin_bot.send_message(msg.chat.id, "❌ خطأ: يرجى إرسال رقم صحيح فقط.")

# ==========================================
# 4. برمجة زر "شحن رصيد" اليدوي
# ==========================================
@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("admin_charge_"))
def admin_charge_prompt(call):
    user_id = int(call.data.replace("admin_charge_", ""))
    admin_bot.answer_callback_query(call.id)
    res = admin_bot.send_message(call.message.chat.id, f"💰 **شحن رصيد:**\nأرسل الآن المبلغ المراد إضافته لحساب العميل (أرقام فقط):", parse_mode="Markdown")
    admin_bot.register_next_step_handler(res, process_admin_recharge, user_id)

def process_admin_recharge(msg, user_id):
    try:
        amount = float(msg.text.strip())
        if amount <= 0:
            return admin_bot.send_message(msg.chat.id, "❌ خطأ: المبلغ يجب أن يكون أكبر من صفر.")
        
        users.update_one({"_id": {"$in": [user_id, str(user_id)]}}, {"$inc": {"balance": amount}})
        user = users.find_one({"_id": {"$in": [user_id, str(user_id)]}})
        new_balance = user.get('balance', 0.0)
        
        admin_bot.send_message(msg.chat.id, f"✅ تم شحن `{amount:.2f}` د.ل بنجاح.\n💰 الرصيد الحالي: `{new_balance:.2f}` د.ل", parse_mode="Markdown")
        try: customer_bot.send_message(user_id, f"💰 **إشعار شحن رصيد:**\nتم إضافة مبلغ `{amount:.2f}` د.ل إلى حسابك.\n💳 رصيدك الحالي هو: `{new_balance:.2f}` د.ل", parse_mode="Markdown")
        except: pass
        show_user_control_card(msg.chat.id, user_id)
    except ValueError:
        admin_bot.send_message(msg.chat.id, "❌ خطأ: يرجى إدخال أرقام فقط.")

# ==========================================
# 5. برمجة زر "تغيير المستوى" 
# ==========================================
@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("change_level_"))
def admin_change_level_menu(call):
    user_id = int(call.data.replace("change_level_", ""))
    kb = types.InlineKeyboardMarkup(row_width=1)
    kb.add(
        types.InlineKeyboardButton("⭐ مستوى 1: قطاعي عادي", callback_data=f"set_level_{user_id}_1"),
        types.InlineKeyboardButton("🥈 مستوى 2: موزع", callback_data=f"set_level_{user_id}_2"),
        types.InlineKeyboardButton("🥇 مستوى 3: جملة", callback_data=f"set_level_{user_id}_3"),
        types.InlineKeyboardButton("🔙 عودة للبطاقة", callback_data=f"back_to_card_{user_id}")
    )
    try: admin_bot.edit_message_text("🏷️ **اختيار مستوى العميل الجديد:**\nسيتم تغيير أسعار المتجر بناءً على اختيارك.", call.message.chat.id, call.message.message_id, reply_markup=kb, parse_mode="Markdown")
    except: pass

@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("set_level_"))
def admin_execute_level_change(call):
    data = call.data.replace("set_level_", "").split("_")
    user_id = int(data[0])
    new_level = int(data[1])
    
    users.update_one({"_id": {"$in": [user_id, str(user_id)]}}, {"$set": {"level": new_level}})
    level_names = {1: "قطاعي عادي", 2: "موزع", 3: "جملة"}
    admin_bot.answer_callback_query(call.id, f"✅ تم تغيير المستوى إلى {level_names[new_level]}")
    
    try: customer_bot.send_message(user_id, f"🎊 **بشرى سارة!**\nتمت ترقية حسابك إلى مستوى (**{level_names[new_level]}**).\nستلاحظ الآن تغيير الأسعار في المتجر تلقائياً.", parse_mode="Markdown")
    except: pass
    show_user_control_card(call.message.chat.id, user_id, call.message.message_id)

@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("back_to_card_"))
def back_to_card_callback(call):
    user_id = int(call.data.replace("back_to_card_", ""))
    show_user_control_card(call.message.chat.id, user_id, call.message.message_id)

# ==========================================
# 6. أزرار "حظر / فك حظر" العميل
# ==========================================
@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("block_user_"))
def block_user_action(call):
    user_id = int(call.data.replace("block_user_", ""))
    users.update_one({"_id": {"$in": [user_id, str(user_id)]}}, {"$set": {"status": "blocked"}})
    admin_bot.answer_callback_query(call.id, "🚫 تم حظر العميل بنجاح.", show_alert=True)
    show_user_control_card(call.message.chat.id, user_id, call.message.message_id)

@admin_bot.callback_query_handler(func=lambda call: call.data.startswith("unblock_user_"))
def unblock_user_action(call):
    user_id = int(call.data.replace("unblock_user_", ""))
    users.update_one({"_id": {"$in": [user_id, str(user_id)]}}, {"$set": {"status": "active"}})
    admin_bot.answer_callback_query(call.id, "✅ تم فك الحظر.", show_alert=True)
    show_user_control_card(call.message.chat.id, user_id, call.message.message_id)

# ==========================================
# 7. المحرك الرئيسي لاستخراج تقرير الإكسيل
# ==========================================
@admin_bot.message_handler(func=lambda m: m.text == "📂 قائمة المشتركين")
def export_subscribers_excel(msg):
    uid = msg.chat.id
    status_msg = admin_bot.send_message(uid, "⏳ **جاري إنشاء التقرير الرسمي لشركة الأهرام...**\nيرجى الانتظار، يتم الآن تنسيق البيانات احترافياً.", parse_mode="Markdown")

    try:
        all_customers = list(users.find({}))
        if not all_customers:
            return admin_bot.edit_message_text("❌ لا يوجد مستخدمين مسجلين حالياً.", uid, status_msg.message_id)

        output = io.BytesIO()
        today_date = datetime.datetime.now().strftime("%Y-%m-%d")
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for customer in all_customers:
                c_id = customer.get("_id")
                c_name = customer.get("name", "عميل_بدون_اسم")
                
                safe_sheet_name = "".join([c for c in str(c_name) if c.isalnum() or c in " "])[:20]
                sheet_name = f"{safe_sheet_name}_{str(c_id)[-4:]}"

                profile_df = pd.DataFrame({
                    "الاسم الكامل": [customer.get("name", "-")],
                    "رقم الهاتف": [customer.get("phone", "-")],
                    "رقم الـ ID": [c_id],
                    "تاريخ التسجيل": [customer.get("reg_date", "غير متوفر")],
                    "الرصيد": [f"{customer.get('balance', 0.0):.2f} د.ل"],
                    "المستوى": [customer.get("level", 1)]
                })

                user_purchases = list(transactions.find({"user_id": c_id}))
                df_purchases = pd.DataFrame([{
                    "المنتج": tx.get("product", "-"),
                    "الفاتورة": tx.get("order_id", "-"),
                    "التاريخ": tx.get("date").strftime("%Y-%m-%d") if hasattr(tx.get("date"), "strftime") else "-",
                    "القيمة": f"{tx.get('total_price', 0.0):.2f} د.ل"
                } for tx in user_purchases]) if user_purchases else pd.DataFrame([{"ملاحظة": "لا توجد مشتريات"}])

                user_complaints = list(db.complaints.find({"user_id": c_id}))
                df_complaints = pd.DataFrame([{
                    "الشكوى": comp.get("reason", "-"),
                    "الحالة": "تم الحل ✅" if comp.get("status") in ["resolved", "resolved_resend", "resolved_replacement", "solved"] else ("تم الإرجاع 💸" if "refund" in str(comp.get("status")) else "انتظار ⏳"),
                    "الرد": comp.get("admin_reply", "-")
                } for comp in user_complaints]) if user_complaints else pd.DataFrame([{"ملاحظة": "لا توجد شكاوي"}])

                profile_df.to_excel(writer, sheet_name=sheet_name, startrow=4, index=False)
                p_end = 4 + len(profile_df) + 2
                df_purchases.to_excel(writer, sheet_name=sheet_name, startrow=p_end, index=False)
                c_start = p_end + len(df_purchases) + 2
                df_complaints.to_excel(writer, sheet_name=sheet_name, startrow=c_start, index=False)

                ws = writer.sheets[sheet_name]
                ws.sheet_view.rightToLeft = True
                
                brand_blue = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
                light_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                gold_accent = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                
                ws.merge_cells('A1:F1')
                title_cell = ws['A1']
                title_cell.value = "شركة الأهرام للاتصالات والتقنية"
                title_cell.fill = brand_blue
                title_cell.font = Font(size=22, bold=True, color="FFFFFF")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 40

                ws.merge_cells('A2:F2')
                date_cell = ws['A2']
                date_cell.value = f"تاريخ استخراج التقرير: {today_date}"
                date_cell.font = Font(size=12, italic=True, color="595959")
                date_cell.alignment = Alignment(horizontal="center")

                def format_header(row_num, text, color_fill):
                    ws.merge_cells(f'A{row_num}:F{row_num}')
                    cell = ws.cell(row=row_num, column=1)
                    cell.value = text
                    cell.fill = color_fill
                    cell.font = Font(bold=True, size=14)
                    cell.alignment = Alignment(horizontal="right")

                format_header(4, "📋 بيانات العميل الأساسية:", light_blue)
                format_header(p_end, "🛒 سجل المشتريات المالي:", light_blue)
                format_header(c_start, "📩 سجل الشكاوي والدعم الفني:", light_blue)

                for row in [5, p_end+1, c_start+1]:
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        if cell.value:
                            cell.fill = gold_accent
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")

                for i in range(1, 7):
                    ws.column_dimensions[get_column_letter(i)].width = 25

        output.seek(0)
        file_name = f"AlAhram_Report_{today_date}.xlsx"
        
        admin_bot.send_document(
            uid, output, visible_file_name=file_name,
            caption=f"✅ **تم توليد تقرير شركة الأهرام بنجاح**\n📂 الملف منظم وجاهز للطباعة أو المراجعة.",
            parse_mode="Markdown"
        )
        admin_bot.delete_message(uid, status_msg.message_id)

    except Exception as e:
        print(f"🚨 خطأ: {e}")
        admin_bot.send_message(uid, "❌ حدث خطأ أثناء تنسيق الملف.")
        
# ==========================================
# 8. البحث السريع عن عميل لضبط الليمت من القائمة الرئيسية
# ==========================================
@admin_bot.message_handler(func=lambda m: m.text == "🛑 ضبط ليمت عميل")
def admin_limit_main_menu(msg):
    uid = msg.chat.id
    res = admin_bot.send_message(uid, "🛑 **ضبط ليمت عميل (السحب بالسالب):**\n\nأرسل الآن **الآيدي (ID)** الخاص بالعميل الذي تريد تعديل حسابه:")
    admin_bot.register_next_step_handler(res, process_limit_user_id_search)

def process_limit_user_id_search(msg):
    uid = msg.chat.id
    try:
        user_id = int(msg.text.strip())
        
        # البحث المرن للتأكد من وجود العميل
        user = users.find_one({"_id": {"$in": [user_id, str(user_id)]}})
        
        if not user:
            return admin_bot.send_message(uid, "❌ العميل غير موجود في قاعدة البيانات. تأكد من صحة الـ ID.")
        
        # 🌟 سحر الكود: استدعاء بطاقة العميل التي برمجناها سابقاً وبها زر الليمت!
        show_user_control_card(uid, user_id)
        
    except ValueError:
        admin_bot.send_message(uid, "❌ خطأ: يرجى إرسال الـ ID كأرقام فقط.")
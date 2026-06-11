import datetime
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ==========================================
# دالة توليد الفاتورة بصيغة Excel (الطريقة الأكثر استقراراً)
# ==========================================
def send_recharge_invoice(uid, amount, user_name, user_phone, customer_bot):
    try:
        invoice_no = f"REC-{datetime.datetime.now().strftime('%y%m%d%H%M')}"
        current_time = datetime.datetime.now().strftime("%Y-%m-%d | %I:%M %p")

        # 1. إنشاء ملف الإكسل
        wb = Workbook()
        ws = wb.active
        ws.title = "إيصال شحن"
        ws.sheet_view.rightToLeft = True # من اليمين لليسار
        
        # 2. التنسيقات والألوان (أزرق داكن وذهبي)
        navy_fill = PatternFill(start_color="1A3A5F", end_color="1A3A5F", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=14)
        blue_bold_font = Font(color="1A3A5F", bold=True, size=12)
        black_bold_font = Font(bold=True, size=12)
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # 3. رسم ترويسة الفاتورة
        ws.merge_cells("A1:B1")
        ws["A1"] = "🏢 شركة الأهرام للاتصالات والتقنية"
        ws["A1"].fill = navy_fill
        ws["A1"].font = white_font
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:B2")
        ws["A2"] = "🧾 إيصال تعبئة كارت"
        ws["A2"].font = blue_bold_font
        ws["A2"].alignment = align_center
        ws.row_dimensions[2].height = 25

        # 4. إدراج بيانات العميل
        details = [
            ("اسم العميل:", user_name),
            ("رقم الهاتف:", user_phone),
            ("رقم الإيصال:", f"#{invoice_no}"),
            ("التاريخ والوقت:", current_time),
            ("القيمة المشحونة إجمالاً:", f"{float(amount):.2f} د.ل")
        ]

        row_start = 4
        for i, (label, value) in enumerate(details, row_start):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = blue_bold_font
            ws[f"A{i}"].alignment = align_right
            
            ws[f"B{i}"] = value
            ws[f"B{i}"].font = black_bold_font
            ws[f"B{i}"].alignment = align_center
            ws.row_dimensions[i].height = 22

        # تظبيط عرض الأعمدة لتبدو كجدول أنيق
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35

        # 5. الحفظ باسم مميز
        file_name = f"Receipt_{uid}_{invoice_no}.xlsx"
        wb.save(file_name)
        
        print(f"⏳ [تتبع]: تم حفظ ملف الإكسل {file_name}، جاري الإرسال...")
        time.sleep(1) # إعطاء الويندوز ثانية ليغلق الملف بعد حفظه

        # 6. إرسال الملف لتليجرام ثم حذفه
        if os.path.exists(file_name):
            with open(file_name, "rb") as doc:
                customer_bot.send_document(
                    uid, 
                    doc, 
                    caption=f"✅ تم شحن حسابك بقيمة {amount:.2f} د.ل\n🧾 مرفق إيصال الشحن الرسمي."
                )
            print("✅ [نجاح]: تم إرسال فاتورة الإكسل للعميل بنجاح.")
            time.sleep(2)
            os.remove(file_name)
        else:
            print("🚨 [خطأ]: ملف الإكسل لم يُحفظ على الهارد ديسك!")

    except Exception as e:
        print(f"🚨 خطأ فادح في فاتورة الإكسل: {e}")
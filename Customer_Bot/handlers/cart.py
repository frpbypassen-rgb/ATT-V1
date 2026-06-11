import datetime
import pandas as pd
from telebot import types
from config import customer_bot, users, cart_db, transactions, stock, admin_bot, ADMIN_ID, counters, invoices, db, ADMIN_IDS
import os
import traceback
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

@customer_bot.message_handler(func=lambda m: m.text and "سلة" in m.text)
def view_cart(msg):
    uid = msg.chat.id
    user_cart = list(cart_db.find({"user_id": uid}))
    
    if not user_cart:
        return customer_bot.send_message(uid, "🛒 سلة المشتريات الخاصة بك فارغة حالياً.")

    total_price = sum(item['total_price'] for item in user_cart)
    
    text = "🛒 **سلة المشتريات الخاصة بك:**\n━━━━━━━━━━━━━━━\n"
    for i, item in enumerate(user_cart, 1):
        text += f"{i}. {item['product']} (الكمية: {item['qty']}) - {item['total_price']} د.ل\n"
    
    text += f"━━━━━━━━━━━━━━━\n💰 **الإجمالي المطلوب:** `{total_price:.2f}` د.ل"

    kb = types.InlineKeyboardMarkup(row_width=1)
    kb.add(
        types.InlineKeyboardButton("💳 إتمام الطلب والدفع", callback_data="checkout_cart"),
        types.InlineKeyboardButton("🗑️ إفراغ السلة", callback_data="clear_cart")
    )
    customer_bot.send_message(uid, text, reply_markup=kb, parse_mode="Markdown")

@customer_bot.callback_query_handler(func=lambda call: call.data == "clear_cart")
def clear_cart_action(call):
    uid = call.message.chat.id
    cart_db.delete_many({"user_id": uid})
    customer_bot.edit_message_text("🗑️ تم إفراغ سلة المشتريات بنجاح.", uid, call.message.message_id)

@customer_bot.callback_query_handler(func=lambda call: call.data == "checkout_cart")
def checkout_process(call):
    uid = call.message.chat.id
    user = users.find_one({"_id": uid})
    user_cart = list(cart_db.find({"user_id": uid}))
    
    # سحب اسم العميل من قاعدة البيانات
    user_name = user.get("name", "عميل غير معروف")

    if not user_cart:
        return customer_bot.answer_callback_query(call.id, "❌ السلة فارغة!", show_alert=True)

    total_price = sum(item['total_price'] for item in user_cart)

    current_balance = user.get("balance", 0.0)
    credit_limit = user.get("credit_limit", 0.0)
    
    if (current_balance + credit_limit) < total_price:
        return customer_bot.answer_callback_query(
            call.id, 
            f"❌ رصيدك ({current_balance:.2f}) وحد الائتمان ({credit_limit:.2f}) لا يكفيان لإتمام العملية.", 
            show_alert=True
        )

    for item in user_cart:
        available_qty = stock.count_documents({"product": item['product'], "sold": False})
        if available_qty < item['qty']:
            return customer_bot.answer_callback_query(
                call.id, 
                f"❌ عذراً، المنتج '{item['product']}' غير متوفر بالكمية المطلوبة. المتاح حالياً {available_qty} كارت فقط.\nيرجى إفراغ السلة وتعديل طلبك.", 
                show_alert=True
            )

    try:
        customer_bot.edit_message_text("⏳ جاري إتمام الطلب وأرشفة الفاتورة...", uid, call.message.message_id)
        
        now_date = datetime.datetime.now()
        yymm_str = now_date.strftime('%y%m')
        
        seq_doc = counters.find_one_and_update(
            {"_id": f"invoice_seq_{yymm_str}"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True
        )
        
        sequence_number = str(seq_doc["seq"]).zfill(3) 
        invoice_id = f"ATT-{yymm_str}-{sequence_number}"

        new_balance = current_balance - total_price
        users.update_one({"_id": uid}, {"$set": {"balance": new_balance}})

        purchased_cards = []
        completed_orders = []

        for item in user_cart:
            cards = list(stock.find({"product": item['product'], "sold": False}).limit(item['qty']))
            c_ids = [c["_id"] for c in cards]
            
            stock.update_many({"_id": {"$in": c_ids}}, {
                "$set": {"sold": True, "sold_to": uid, "sold_date": now_date, "order_id": invoice_id}
            })
            
            purchased_cards.extend(cards)

            completed_orders.append({
                "order_id": invoice_id,
                "user_id": uid,
                "product": item['product'],
                "qty": item['qty'],
                "total_price": item['total_price'],
                "date": now_date
            })
        
        transactions.insert_many(completed_orders)
        cart_db.delete_many({"user_id": uid})

        invoices.insert_one({
            "invoice_id": invoice_id,
            "user_id": uid,
            "total_amount": total_price,
            "items_count": len(user_cart),
            "date": now_date
        })

        file_path = f"Invoice_{invoice_id}.xlsx"
        
        # تمرير اسم العميل إلى دالة صناعة الفاتورة
        generate_cart_invoice(purchased_cards, completed_orders, file_path, invoice_id, user_name)

        # 1. إرسال الفاتورة للعميل
        with open(file_path, "rb") as doc:
            caption = (
                f"✅ **تم تنفيذ الطلبات بنجاح!**\n"
                f"🧾 **رقم الفاتورة:** `{invoice_id}`\n"
                f"💰 **المبلغ المخصوم:** `{total_price:.2f}` د.ل\n"
                f"💳 **رصيدك المتبقي:** `{new_balance:.2f}` د.ل\n\n"
                f"📂 *(تجد ملخص الطلبية والأكواد داخل الفاتورة المرفقة)*"
            )
            customer_bot.send_document(uid, doc, caption=caption, visible_file_name=f"{invoice_id}.xlsx", parse_mode="Markdown")
        
        # ==========================================
        # 🌟 التحديث: توزيع الإشعار والفاتورة على كافة المشرفين
        # ==========================================
        all_admins = [str(ADMIN_ID)]
        try:
            if isinstance(ADMIN_IDS, list):
                all_admins.extend([str(a) for a in ADMIN_IDS])
        except: pass
        
        try:
            for doc_admin in db.admins.find():
                all_admins.append(str(doc_admin["_id"]))
        except: pass

        caption_admin = f"🚨 **مبيعات السلة!**\nالعميل: `{user_name}` (`{uid}`)\nالفاتورة: `{invoice_id}`\nالقيمة: `{total_price:.2f}` د.ل"
        
        # إرسال لكل مشرف بدون تكرار
        for admin in set(all_admins):
            try:
                # يجب فتح الملف مجدداً لكل إرسال لضمان عدم حدوث خطأ في المؤشر (File Pointer)
                with open(file_path, "rb") as doc_admin:
                    admin_bot.send_document(
                        admin, doc_admin,
                        caption=caption_admin,
                        visible_file_name=f"Copy_{invoice_id}.xlsx"
                    )
            except Exception as e: 
                print(f"🚨 تعذر الإرسال للمشرف {admin}: {e}")

        # أرشفة الملف
        archive_dir = os.path.join(os.getcwd(), "Invoices_Archive")
        os.makedirs(archive_dir, exist_ok=True)  
        
        final_archive_path = os.path.join(archive_dir, f"{invoice_id}.xlsx")
        os.replace(file_path, final_archive_path)

    except Exception as e:
        print(f"🚨 Error: {traceback.format_exc()}")
        customer_bot.send_message(uid, "❌ حدث خطأ أثناء إصدار الفاتورة. يرجى مراجعة الدعم الفني.")

# ==========================================
# دالة بناء الفاتورة (الملخص + رسالة الشكر + حماية النصوص)
# ==========================================
def generate_cart_invoice(cards, summary_orders, path, invoice_id, user_name):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        
        summary_list = []
        for t in summary_orders:
            qty = t.get('qty', 1)
            total = t.get('total_price', 0)
            unit_price = total / qty if qty > 0 else 0
            
            summary_list.append({
                'اسم المنتج': t.get('product'),
                'العدد': qty,
                'السعر': unit_price,
                'المجموع': total
            })
            
        df_sum = pd.DataFrame(summary_list)
        df_sum.insert(0, 'ت', range(1, 1 + len(df_sum)))
        df_sum.to_excel(writer, sheet_name="📊 ملخص الفاتورة", index=False, startrow=6)
        
        products = set([c.get('product', 'Unknown') for c in cards])
        for prod in products:
            prod_cards = [c for c in cards if c.get('product') == prod]
            df = pd.DataFrame(prod_cards)
            
            expected_cols = ['product', 'code', 'serial', 'pin', 'op_code']
            for col in expected_cols:
                if col not in df.columns: df[col] = pd.NA
                    
            # 🌟 الحماية 1: تحويل السيريال والأكواد إلى نصوص وإزالة .0
            for col in ['code', 'pin', 'serial', 'op_code']:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)
                    df[col] = df[col].replace(['nan', 'None', 'NaN', ''], '-')

            df = df[expected_cols]
            df.rename(columns={'product': 'المنتج', 'code': 'الكود', 'serial': 'السيريال', 'pin': 'الرقم السري', 'op_code': 'أوبريشن'}, inplace=True)
            df.insert(0, 'ت', range(1, 1 + len(df)))

            safe_sheet_name = "".join([c for c in prod if c.isalnum() or c in " _-"])[:30]
            if not safe_sheet_name: safe_sheet_name = "مشتريات"
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False, startrow=6)

        workbook = writer.book
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        password_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        footer_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=14)
        bold_font = Font(bold=True, size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal="center", vertical="center")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.sheet_view.rightToLeft = True
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A5
            
            num_cols = worksheet.max_column
            last_col_letter = get_column_letter(num_cols)

            worksheet.merge_cells(f'A1:{last_col_letter}1')
            worksheet["A1"] = "🏢 شركة الأهرام للاتصالات والتقنية"
            worksheet["A1"].font = white_font
            worksheet["A1"].fill = header_fill
            worksheet["A1"].alignment = align
            worksheet.row_dimensions[1].height = 30

            worksheet.merge_cells(f'A2:{last_col_letter}2')
            worksheet["A2"] = f"تفاصيل الفاتورة"
            worksheet["A2"].font = Font(bold=True, size=12)
            worksheet["A2"].alignment = align

            worksheet.merge_cells(f'A3:{last_col_letter}3')
            worksheet["A3"] = f"رقم الفاتورة: {invoice_id}"
            worksheet["A3"].alignment = align

            worksheet.merge_cells(f'A4:{last_col_letter}4')
            worksheet["A4"] = f"اسم العميل: {user_name}"
            worksheet["A4"].alignment = align
            worksheet["A4"].font = Font(bold=True, color="1F4E78", size=12)

            worksheet.merge_cells(f'A5:{last_col_letter}5')
            worksheet["A5"] = f"التاريخ: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
            worksheet["A5"].alignment = align

            for col_num in range(1, num_cols + 1):
                cell = worksheet.cell(row=7, column=col_num)
                cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True); cell.alignment = align; cell.border = thin_border

            for row_idx in range(8, worksheet.max_row + 1):
                for col_idx in range(1, num_cols + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = align; cell.border = thin_border; cell.font = bold_font
                    
                    # 🌟 الحماية 2: إجبار الإكسيل على قراءة الخلية كنص (Text)
                    cell.number_format = '@'
                    
                    if cell.value is not None and str(cell.value) != '-':
                        cell.value = str(cell.value)
                    
                    if "ملخص" not in sheet_name and worksheet.cell(row=7, column=col_idx).value == 'الرقم السري':
                        cell.fill = password_fill

            for i, col in enumerate(worksheet.columns, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = 20

            if "ملخص" in sheet_name:
                last_r = worksheet.max_row + 2
                grand_total = sum(d['المجموع'] for d in summary_list)
                total_qty = sum(d['الكمية'] if 'الكمية' in d else d['العدد'] for d in summary_list)
                
                worksheet.cell(row=last_r, column=2, value="العدد الكلي:").fill = footer_fill
                worksheet.cell(row=last_r, column=2).font = bold_font; worksheet.cell(row=last_r, column=2).alignment = align
                
                worksheet.cell(row=last_r, column=3, value=total_qty).fill = footer_fill
                worksheet.cell(row=last_r, column=3).font = bold_font; worksheet.cell(row=last_r, column=3).alignment = align
                
                worksheet.cell(row=last_r, column=4, value="المبلغ الكلي:").fill = footer_fill
                worksheet.cell(row=last_r, column=4).font = bold_font; worksheet.cell(row=last_r, column=4).alignment = align
                
                worksheet.cell(row=last_r, column=5, value=f"{grand_total:.2f} د.ل").fill = footer_fill
                worksheet.cell(row=last_r, column=5).font = bold_font; worksheet.cell(row=last_r, column=5).alignment = align
                
                thanks_row = last_r + 2
                worksheet.merge_cells(start_row=thanks_row, start_column=1, end_row=thanks_row+1, end_column=num_cols)
                thanks_cell = worksheet.cell(row=thanks_row, column=1)
                thanks_cell.value = "💖 شكراً لثقتكم وتسوقكم من شركة الأهرام للاتصالات والتقنية.. نتمنى لكم يوماً سعيداً!"
                thanks_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thanks_cell.font = Font(color="1F4E78", bold=True, size=13, italic=True)
                thanks_cell.fill = PatternFill(start_color="EAEEF7", end_color="EAEEF7", fill_type="solid")
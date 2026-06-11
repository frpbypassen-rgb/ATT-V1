import datetime
import pandas as pd
from telebot import types
from config import customer_bot, transactions, users
import os
import traceback
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

print("✅ تم قراءة ملف سجل المشتريات بنجاح!")

# ==========================================
# 1. الاستماع لزر سجل المشتريات
# ==========================================
@customer_bot.message_handler(func=lambda m: m.text and "سجل" in m.text)
def history_menu(msg):
    try:
        print("🚨 [رادار السجل]: تم استقبال طلب عرض سجل المشتريات.")
        uid = msg.chat.id
        user = users.find_one({"_id": uid})
        
        if not user or user.get("status") != "active":
            return customer_bot.send_message(uid, "⚠️ عذراً، يجب تفعيل حسابك أولاً.")

        kb = types.InlineKeyboardMarkup(row_width=2)
        kb.add(
            types.InlineKeyboardButton("📅 سجل اليوم", callback_data="hist_today"),
            types.InlineKeyboardButton("🗓️ سجل الشهر", callback_data="hist_month")
        )
        
        customer_bot.send_message(uid, "📊 **سجل مشتريات شركة الأهرام**\nاختر الفترة المطلوبة للتقرير:", reply_markup=kb, parse_mode="Markdown")
    except Exception as e:
        print(f"🚨 خطأ في قائمة السجل: {e}")

# ==========================================
# 2. معالجة طلب استخراج التقرير
# ==========================================
@customer_bot.callback_query_handler(func=lambda call: call.data.startswith("hist_"))
def generate_history_excel(call):
    uid = call.message.chat.id
    period = call.data.split("_")[1]
    
    print(f"⏳ جاري استخراج السجل ({period}) للعميل {uid}...")
    
    try:
        now = datetime.datetime.now()
        if period == "today":
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            file_label = "اليومي"
        else:
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            file_label = "الشهري"

        # جلب البيانات من الداتابيز
        query = {"user_id": uid, "date": {"$gte": start_date}}
        data = list(transactions.find(query).sort("date", -1))

        if not data:
            return customer_bot.answer_callback_query(call.id, f"❌ لا توجد عمليات في السجل {file_label}.", show_alert=True)

        customer_bot.answer_callback_query(call.id, "⏳ جاري توليد ملف الإكسل المنسق...")
        
        # إنشاء الملف
        file_path = f"Report_{uid}_{period}.xlsx"
        create_excel_report(data, file_path, file_label)

        # إرسال الملف للعميل
        with open(file_path, "rb") as doc:
            caption = (
                f"✅ سجل المشتريات {file_label}\n"
                f"🏢 شركة الأهرام للاتصالت والتقنية\n"
                f"📅 الفترة: {start_date.strftime('%Y-%m-%d')} إلى {now.strftime('%Y-%m-%d')}"
            )
            customer_bot.send_document(
                uid, doc, 
                caption=caption, 
                visible_file_name=f"AlAhram_Report_{file_label}.xlsx"
            )
        
        os.remove(file_path) # حذف الملف المؤقت
        print(f"✅ تم إرسال التقرير {file_label} بنجاح.")

    except Exception as e:
        print(f"🚨 فشل استخراج الإكسل! السبب:\n{traceback.format_exc()}")
        customer_bot.answer_callback_query(call.id, "❌ حدث خطأ فني أثناء إعداد الملف.", show_alert=True)

# ==========================================
# 3. دالة بناء ملف الإكسل مع المجموع الكلي
# ==========================================
def create_excel_report(data, path, label):
    report_list = []
    total_sum = 0.0 # لحساب المجموع الكلي
    
    for i, op in enumerate(data, 1):
        price = float(op.get('total_price', 0))
        total_sum += price
        report_list.append({
            "ت": i,
            "رقم العملية": op.get("order_id", "---"),
            "المنتج": op.get("product", "---"),
            "القيمة (د.ل)": price, # نضعه كرقم ليتمكن الإكسل من التعامل معه
            "التاريخ": op.get("date").strftime("%Y-%m-%d %H:%M")
        })

    df = pd.DataFrame(report_list)
    
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=4)
        
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A5
        worksheet.sheet_view.rightToLeft = True # التنسيق من اليمين لليسار

        # --- تعريف الألوان والتنسيقات ---
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # أزرق غامق
        table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # رمادي فاتح
        white_font = Font(color="FFFFFF", bold=True, size=14)
        black_bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # 1. تنسيق العناوين العلوية
        worksheet["A1"] = "🏢 شركة الأهرام للاتصالت والتقنية"
        worksheet["A1"].font = white_font
        worksheet["A1"].fill = header_fill
        worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet.merge_cells("A1:E1")
        worksheet.row_dimensions[1].height = 30

        worksheet["A2"] = f"تقرير المشتريات {label}"
        worksheet["A2"].font = Font(bold=True, size=12)
        worksheet["A2"].alignment = Alignment(horizontal="center")
        worksheet.merge_cells("A2:E2")

        worksheet["A3"] = f"تاريخ الاستخراج: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
        worksheet["A3"].alignment = Alignment(horizontal="center")
        worksheet.merge_cells("A3:E3")

        # 2. تنسيق ترويسة الجدول (السطر 5)
        for col_num in range(1, 6):
            cell = worksheet.cell(row=5, column=col_num)
            cell.fill = table_header_fill
            cell.font = black_bold_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 3. تنسيق خلايا البيانات (تبدأ من السطر 6)
        last_data_row = len(data) + 5
        for row in worksheet.iter_rows(min_row=6, max_row=last_data_row, min_col=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
                if cell.column == 4: # عمود القيمة
                    cell.number_format = '#,##0.00'

        # 🌟 4. إضافة سطر المجموع الكلي في الأسفل
        total_row_idx = last_data_row + 1
        
        # دمج أول 3 خلايا لكلمة "المجموع الكلي"
        worksheet.merge_cells(f"A{total_row_idx}:C{total_row_idx}")
        total_label_cell = worksheet[f"A{total_row_idx}"]
        total_label_cell.value = "المجموع الكلي"
        total_label_cell.font = black_bold_font
        total_label_cell.alignment = Alignment(horizontal="center")
        total_label_cell.fill = table_header_fill
        total_label_cell.border = thin_border
        
        # وضع قيمة المجموع في خلية القيمة
        total_value_cell = worksheet[f"D{total_row_idx}"]
        total_value_cell.value = total_sum
        total_value_cell.font = black_bold_font
        total_value_cell.alignment = Alignment(horizontal="center")
        total_value_cell.border = thin_border
        total_value_cell.number_format = '#,##0.00'
        
        # إضافة حدود للخلية الأخيرة الفارغة في سطر المجموع ليكون التنسيق متكاملاً
        worksheet[f"E{total_row_idx}"].border = thin_border

        # 5. ضبط عرض الأعمدة تلقائياً
        for i, col in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(i)
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except: pass
            worksheet.column_dimensions[column_letter].width = max_length + 5
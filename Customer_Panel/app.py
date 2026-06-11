import sys
import os
import datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash

# ربط موقع العميل بقاعدة البيانات المشتركة
root_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.insert(0, root_path)

try:
    from config import db, users, stock, transactions
except ImportError:
    print("🚨 خطأ: لم يتم العثور على ملف config.py")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, 'templates')
STATIC_DIR = os.path.join(BASE_DIR, 'static')

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.secret_key = "AlAhram_Customer_Portal_2026_Secure"

@app.route('/login', methods=['GET', 'POST'])
def login():
    if session.get('customer_logged_in'):
        return redirect(url_for('index'))

    if request.method == 'POST':
        tg_id = request.form.get('tg_id', '').strip()
        password = request.form.get('password', '').strip()

        if not password.isdigit() or len(password) != 6:
            flash("كلمة المرور يجب أن تتكون من 6 أرقام فقط!", "danger")
            return redirect(url_for('login'))

        try: tg_id_int = int(tg_id)
        except: tg_id_int = tg_id

        user = users.find_one({
            "$or": [{"_id": tg_id_int}, {"_id": str(tg_id)}],
            "web_password": password
        })

        if user:
            session['customer_logged_in'] = True
            session['customer_id'] = user.get('_id')
            session['customer_name'] = user.get('name') or user.get('first_name') or 'عميلنا العزيز'
            return redirect(url_for('index'))
        else:
            flash("بيانات الدخول غير صحيحة! تأكد من رقم التليجرام وكلمة المرور.", "danger")

    return render_template('login.html')

@app.route('/')
def index():
    if not session.get('customer_logged_in'):
        return redirect(url_for('login'))
    
    user_id = session.get('customer_id')
    user = users.find_one({"_id": user_id})
    
    if not user:
        session.clear()
        return redirect(url_for('login'))
        
    # جلب وتنسيق الرصيد
    balance = "{:.2f}".format(float(user.get('balance', 0)))
    
    # جلب آخر الفواتير والمشتريات
    user_trans = list(transactions.find({"user_id": {"$in": [user_id, str(user_id)]}}).sort("date", -1))
    
    total_spent = 0
    recent_orders = []
    seen_orders = set()
    
    for t in user_trans:
        oid = t.get('order_id')
        price = t.get('total_price', 0)
        total_spent += price
        
        if oid and oid not in seen_orders:
            seen_orders.add(oid)
            # إضافة آخر 5 طلبات فقط للعرض السريع في الرئيسية
            if len(recent_orders) < 5:
                d_obj = t.get('date')
                date_str = d_obj.strftime('%Y-%m-%d') if isinstance(d_obj, datetime.datetime) else "---"
                recent_orders.append({
                    "order_id": oid,
                    "date": date_str,
                    "price": "{:.2f}".format(float(price))
                })
                
    formatted_total = "{:.2f}".format(float(total_spent))

    return render_template('index.html', 
                           customer_name=session.get('customer_name'), 
                           balance=balance, 
                           total_spent=formatted_total,
                           recent_orders=recent_orders)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def get_customer_name(user_id):
    if not user_id: return "مجهول"
    try: uid_int = int(user_id)
    except: uid_int = None
    
    u_data = users.find_one({"$or": [{"_id": uid_int}, {"_id": str(user_id)}]})
    if u_data:
        name = u_data.get('name') or u_data.get('first_name') or u_data.get('username')
        if name: return str(name).strip()
    return str(user_id)

@app.route('/invoices')
def invoices_list():
    if not session.get('customer_logged_in'):
        return redirect(url_for('login'))
        
    user_id = session.get('customer_id')
    user = users.find_one({"_id": user_id})
    if not user:
        return redirect(url_for('login'))
        
    # جلب كافة معاملات العميل
    user_trans = list(transactions.find({"user_id": {"$in": [user_id, str(user_id)]}}).sort("date", -1))
    
    invoices_dict = {}
    for t in user_trans:
        oid = t.get('order_id')
        if not oid: continue
        price = t.get('total_price', 0)
        
        if oid not in invoices_dict:
            cards_count = stock.count_documents({"order_id": oid})
            invoices_dict[oid] = {
                "order_id": oid,
                "total_price": 0,
                "date": t.get('date'),
                "cards_count": cards_count if cards_count > 0 else 1
            }
        invoices_dict[oid]["total_price"] += price
        
    enriched_invoices = []
    for oid, data in invoices_dict.items():
        d_obj = data["date"]
        data["formatted_price"] = "{:.2f}".format(float(data["total_price"]))
        data["date_str"] = d_obj.strftime('%Y-%m-%d') if isinstance(d_obj, datetime.datetime) else "---"
        data["time_str"] = d_obj.strftime('%H:%M:%S') if isinstance(d_obj, datetime.datetime) else "---"
        enriched_invoices.append(data)
        
    enriched_invoices.sort(key=lambda x: x.get("date") if isinstance(x.get("date"), datetime.datetime) else datetime.datetime.min, reverse=True)
    
    balance = "{:.2f}".format(float(user.get('balance', 0)))
    return render_template('invoices.html', invoices=enriched_invoices, balance=balance)

@app.route('/download_invoice/<order_id>')
def download_invoice(order_id):
    if not session.get('customer_logged_in'):
        return redirect(url_for('login'))
        
    user_id = session.get('customer_id')
    
    # تحقق أمني: التأكد من أن هذه الفاتورة تنتمي للعميل الحالي
    all_trans = list(transactions.find({"order_id": order_id, "user_id": {"$in": [user_id, str(user_id)]}}))
    if not all_trans:
        return "غير مسموح لك بتحميل هذه الفاتورة أو أنها غير موجودة.", 403
        
    cards = list(stock.find({"order_id": order_id}))
    customer_name = get_customer_name(user_id)
    order_date = all_trans[0].get('date')
    date_str = order_date.strftime('%Y-%m-%d %H:%M') if isinstance(order_date, datetime.datetime) else "---"

    import io
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import Response

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_list = []
        for t in all_trans:
            qty = t.get('qty', 1)
            total = t.get('total_price', 0)
            unit_price = total / qty if qty > 0 else 0
            summary_list.append({'اسم المنتج': t.get('product'), 'العدد': qty, 'السعر': unit_price, 'المجموع': total})
        df_sum = pd.DataFrame(summary_list)
        df_sum.insert(0, 'ت', range(1, 1 + len(df_sum)))
        df_sum.to_excel(writer, sheet_name="📊 ملخص الفاتورة", index=False, startrow=6)
        
        products = set([c.get('product', 'Unknown') for c in cards])
        for prod in products:
            prod_cards = [c for c in cards if c.get('product') == prod]
            df = pd.DataFrame(prod_cards)
            expected_cols = ['product', 'code', 'serial', 'pin', 'op_code']
            for col in expected_cols:
                if col not in df.columns: df[col] = pd.NA
            for col in ['code', 'pin', 'serial', 'op_code']:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)
                    df[col] = df[col].replace(['nan', 'None', 'NaN', ''], '-')
            df = df[expected_cols]
            df.rename(columns={'product': 'المنتج', 'code': 'الكود', 'serial': 'السيريال', 'pin': 'الرقم السري', 'op_code': 'أوبريشن'}, inplace=True)
            df.insert(0, 'ت', range(1, 1 + len(df)))
            safe_sheet_name = "".join([c for c in prod if c.isalnum() or c in " _-"])[:30]
            if not safe_sheet_name: safe_sheet_name = "مشتريات"
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False, startrow=6)

        workbook = writer.book
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        password_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        footer_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=14)
        bold_font = Font(bold=True, size=11)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal="center", vertical="center")

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.sheet_view.rightToLeft = True
            worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A5
            last_col_letter = get_column_letter(worksheet.max_column)

            worksheet.merge_cells(f'A1:{last_col_letter}1')
            worksheet["A1"] = "🏢 شركة الأهرام للاتصالات والتقنية"
            worksheet["A1"].font, worksheet["A1"].fill, worksheet["A1"].alignment = white_font, header_fill, align
            worksheet.row_dimensions[1].height = 30

            worksheet.merge_cells(f'A2:{last_col_letter}2'); worksheet["A2"] = "تفاصيل الفاتورة"
            worksheet["A2"].font, worksheet["A2"].alignment = Font(bold=True, size=12), align

            worksheet.merge_cells(f'A3:{last_col_letter}3'); worksheet["A3"] = f"رقم الفاتورة: {order_id}"
            worksheet["A3"].alignment = align

            worksheet.merge_cells(f'A4:{last_col_letter}4'); worksheet["A4"] = f"اسم العميل: {customer_name}"
            worksheet["A4"].alignment, worksheet["A4"].font = align, Font(bold=True, color="1F4E78", size=12)

            worksheet.merge_cells(f'A5:{last_col_letter}5'); worksheet["A5"] = f"التاريخ: {date_str}"
            worksheet["A5"].alignment = align

            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=7, column=col_num)
                cell.fill, cell.font, cell.alignment, cell.border = header_fill, Font(color="FFFFFF", bold=True), align, thin_border

            for row_idx in range(8, worksheet.max_row + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment, cell.border, cell.font = align, thin_border, bold_font
                    cell.number_format = '@'
                    if cell.value is not None and str(cell.value) != '-':
                        cell.value = str(cell.value)
                    if "ملخص" not in sheet_name and worksheet.cell(row=7, column=col_idx).value == 'الرقم السري':
                        cell.fill = password_fill

            for i, col in enumerate(worksheet.columns, 1):
                worksheet.column_dimensions[get_column_letter(i)].width = 20

            if "ملخص" in sheet_name:
                last_r = worksheet.max_row + 2
                grand_total = sum(d['المجموع'] for d in summary_list)
                total_qty = sum(d['العدد'] for d in summary_list)
                
                worksheet.cell(row=last_r, column=2, value="العدد الكلي:").fill = footer_fill
                worksheet.cell(row=last_r, column=2).font, worksheet.cell(row=last_r, column=2).alignment = bold_font, align
                worksheet.cell(row=last_r, column=3, value=total_qty).fill = footer_fill
                worksheet.cell(row=last_r, column=3).font, worksheet.cell(row=last_r, column=3).alignment = bold_font, align
                worksheet.cell(row=last_r, column=4, value="المبلغ الكلي:").fill = footer_fill
                worksheet.cell(row=last_r, column=4).font, worksheet.cell(row=last_r, column=4).alignment = bold_font, align
                worksheet.cell(row=last_r, column=5, value=f"{grand_total:.2f} د.ل").fill = footer_fill
                worksheet.cell(row=last_r, column=5).font, worksheet.cell(row=last_r, column=5).alignment = bold_font, align
                
                thanks_row = last_r + 2
                worksheet.merge_cells(start_row=thanks_row, start_column=1, end_row=thanks_row+1, end_column=worksheet.max_column)
                thanks_cell = worksheet.cell(row=thanks_row, column=1)
                thanks_cell.value = "💖 شكراً لثقتكم وتسوقكم من شركة الأهرام للاتصالات والتقنية.. نتمنى لكم يوماً سعيداً!"
                thanks_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thanks_cell.font, thanks_cell.fill = Font(color="1F4E78", bold=True, size=13, italic=True), PatternFill(start_color="EAEEF7", end_color="EAEEF7", fill_type="solid")

    output.seek(0)
    return Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-disposition": f"attachment; filename=Invoice_{order_id}.xlsx"})

if __name__ == '__main__':
    app.run(debug=True, port=5001)
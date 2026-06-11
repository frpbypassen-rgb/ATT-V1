import openpyxl
import csv
import io
import math
import re

NA = None
nan = float('nan')

class ExcelWriter:
    def __init__(self, path_or_buf, engine=None):
        self.path_or_buf = path_or_buf
        self.book = openpyxl.Workbook()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if len(self.book.sheetnames) > 1 and "Sheet" in self.book.sheetnames:
            self.book.remove(self.book["Sheet"])
        
        if isinstance(self.path_or_buf, io.BytesIO):
            self.book.save(self.path_or_buf)
        else:
            self.book.save(self.path_or_buf)

    @property
    def sheets(self):
        return {ws.title: ws for ws in self.book.worksheets}

class StringMethods:
    def __init__(self, series):
        self.series = series

    def replace(self, pat, repl, n=-1, case=None, flags=0, regex=False):
        new_vals = []
        for val in self.series.values:
            val_str = str(val) if val is not None and not (isinstance(val, float) and math.isnan(val)) else ""
            if regex:
                try:
                    val_str = re.sub(pat, repl, val_str)
                except Exception as e:
                    print(f"Regex replace error: {e}")
            else:
                val_str = val_str.replace(pat, repl)
            new_vals.append(val_str)
        return Series(new_vals)

class Series:
    def __init__(self, values):
        self.values = list(values)

    @property
    def str(self):
        return StringMethods(self)

    def astype(self, dtype):
        if dtype == str:
            new_vals = []
            for v in self.values:
                if v is None:
                    new_vals.append("nan")
                elif isinstance(v, float) and math.isnan(v):
                    new_vals.append("nan")
                else:
                    new_vals.append(str(v))
            return Series(new_vals)
        return self

    def replace(self, to_replace, value=None):
        new_vals = []
        for v in self.values:
            val_str = str(v) if v is not None and not (isinstance(v, float) and math.isnan(v)) else ""
            if isinstance(to_replace, list):
                if v in to_replace or val_str in to_replace:
                    new_vals.append(value)
                else:
                    new_vals.append(v)
            else:
                if v == to_replace or val_str == to_replace:
                    new_vals.append(value)
                else:
                    new_vals.append(v)
        return Series(new_vals)

class DataFrame:
    def __init__(self, data=None, columns=None):
        self._data = []
        if data is None:
            pass
        elif isinstance(data, list):
            for item in data:
                if isinstance(item, dict):
                    self._data.append(dict(item))
                else:
                    self._data.append(list(item))
        elif isinstance(data, dict):
            keys = list(data.keys())
            length = len(data[keys[0]]) if keys else 0
            for i in range(length):
                self._data.append({k: data[k][i] for k in keys})
        
        if columns is not None:
            self.columns = list(columns)
        else:
            if self._data and isinstance(self._data[0], dict):
                self.columns = list(self._data[0].keys())
            else:
                self.columns = []

    @property
    def empty(self):
        return len(self._data) == 0

    def insert(self, loc, column, value):
        val_list = list(value)
        while len(self._data) < len(val_list):
            self._data.append({})
        for i, val in enumerate(val_list):
            if isinstance(self._data[i], dict):
                self._data[i][column] = val
        if column not in self.columns:
            self.columns.insert(loc, column)

    def rename(self, columns, inplace=True):
        new_data = []
        for row in self._data:
            if isinstance(row, dict):
                new_row = {}
                for k, v in row.items():
                    new_row[columns.get(k, k)] = v
                new_data.append(new_row)
            else:
                new_data.append(row)
        
        new_cols = [columns.get(c, c) for c in self.columns]
        
        if inplace:
            self._data = new_data
            self.columns = new_cols
            return self
        else:
            return DataFrame(new_data, columns=new_cols)

    def iterrows(self):
        for i, row in enumerate(self._data):
            yield i, row

    def __getitem__(self, item):
        if isinstance(item, list):
            new_data = []
            for row in self._data:
                if isinstance(row, dict):
                    new_data.append({col: row.get(col, None) for col in item})
            return DataFrame(new_data, columns=item)
        else:
            vals = []
            for row in self._data:
                if isinstance(row, dict):
                    vals.append(row.get(item, None))
            return Series(vals)

    def __setitem__(self, key, value):
        if isinstance(value, Series):
            val_list = value.values
        elif isinstance(value, list) or isinstance(value, range):
            val_list = list(value)
        else:
            val_list = [value] * len(self._data)
            
        while len(self._data) < len(val_list):
            self._data.append({})
            
        for i in range(max(len(self._data), len(val_list))):
            if i >= len(self._data):
                self._data.append({})
            val = val_list[i] if i < len(val_list) else None
            if isinstance(self._data[i], dict):
                self._data[i][key] = val
                
        if key not in self.columns:
            self.columns.append(key)

    def to_excel(self, writer, sheet_name="Sheet1", index=False, startrow=0):
        wb = writer.book
        if sheet_name not in wb.sheetnames:
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                ws = wb["Sheet"]
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb[sheet_name]
            
        r = startrow + 1
        for col_idx, col_name in enumerate(self.columns, 1):
            ws.cell(row=r, column=col_idx, value=col_name)
            
        for row_idx, row in enumerate(self._data, r + 1):
            for col_idx, col_name in enumerate(self.columns, 1):
                val = row.get(col_name, None) if isinstance(row, dict) else (row[col_idx-1] if col_idx-1 < len(row) else None)
                if isinstance(val, float) and math.isnan(val):
                    val = None
                ws.cell(row=row_idx, column=col_idx, value=val)

def read_excel(filepath, skiprows=0):
    if isinstance(filepath, io.BytesIO):
        wb = openpyxl.load_workbook(filepath)
    else:
        wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= skiprows:
        return DataFrame()
    
    header_row = rows[skiprows]
    columns = [str(cell) if cell is not None else f"Unnamed: {i}" for i, cell in enumerate(header_row)]
    
    data = []
    for row in rows[skiprows + 1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for col_idx, col_name in enumerate(columns):
            val = row[col_idx] if col_idx < len(row) else None
            row_dict[col_name] = val
        data.append(row_dict)
        
    return DataFrame(data, columns=columns)

def read_csv(filepath):
    data = []
    try:
        with open(filepath, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            columns = reader.fieldnames
            for row in reader:
                data.append(dict(row))
    except:
        with open(filepath, mode='r', encoding='cp1252') as f:
            reader = csv.DictReader(f)
            columns = reader.fieldnames
            for row in reader:
                data.append(dict(row))
                
    return DataFrame(data, columns=columns)
